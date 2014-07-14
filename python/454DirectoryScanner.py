#!/usr/bin/env python3.2

import csv, argparse, sys, datetime, time, random, os.path, shutil, json, xlrd,re
from openpyxl import load_workbook
import openpyxl.cell

#-------------------------------------------------------------------------------	

##########

def handle_454_jr (file_path, mb_lookup):
    #PID (MM/DD/YYYY) junk
    pid_format       = re.compile ('^[\s]*([\w\-]+)[\s]*$')
    
    #print (file_path)
    ngs_records = {}
    try:
        wb = load_workbook (file_path)
        ws = wb.active
        #print (file_path, ws)
        for r in range (1,50):
            try:    
                cells = [ws.cell ('%s%d' %(openpyxl.cell.get_column_letter (k), r)).value for k in range (1,7)]
            
                run_id = None
                p_info = None
            
                for c in range (6):
                    if type (cells[c]) == str:
                        m = pid_format.match (cells[c])
                        if m is not None:
                            try:
                                pid = mb_lookup[m.group (1)]
                                if type (cells[c+1]) == datetime.datetime:
                                    date = cells[c+1].strftime("%Y%m%d")
                                    p_info = (pid, date)
                                if type (cells[c+2]) == int:
                                    run_id = cells[c+2]
                                
                            except (IndexError,KeyError) as e:
                                pass
                            
                                                                    
                if run_id is not None and p_info is not None:
                    if run_id in ngs_records:
                        print (cells)
                        print ("DUPLICATE RUN_ID %d in %s" % (run_id, file_path), file = sys.stderr)
                        return None
                    else:
                        ngs_records[run_id] = p_info
                
            except:
                raise
    except:
        print ("FAILED reading XLSX %s" % file_path, file = sys.stderr)
        #raise
            
    return ngs_records if len (ngs_records) > 0 else None   
    
##########

def handle_GeneCore_submission (file_path, mb_lookup):
    #PID (MM/DD/YYYY) junk
    pid_format       = re.compile ('^([\w\-]+)\s+\(([0-9]{2})/([0-9]{2})/([0-9]{4})\).+$')
    sample_id_format = re.compile ('^Sample\s+\#\s*([0-9]+).*$')
    
    #print (file_path)
    ngs_records = {}
    try:
        wb = load_workbook (file_path)
        ws = wb.active
        #print (file_path, ws)
        for r in range (1,50):
            try:    
                cells = [ws.cell ('%s%d' %(openpyxl.cell.get_column_letter (k), r)).value for k in range (1,7)]
            
                run_id = None
                p_info = None
            
                for c in range (6):
                    if type (cells[c]) == str:
                        m = pid_format.match (cells[c])
                        if m is not None:
                            try:
                                pid = mb_lookup[m.group (1)]
                                date = datetime.datetime (int(m.group (4)), int(m.group (2)), int(m.group(3)))
                                p_info = (pid,date.strftime("%Y%m%d"))
                            except KeyError as e:
                                pass
                        else:
                            #print (cells[c])
                            m = sample_id_format.match (cells[c])
                            #print (m)
                            if m is not None:
                                try:
                                    run_id = int (m.group(1))
                                    if run_id < 1 or run_id > 16:
                                        run_id = None
                                except:
                                    pass
                    else:
                        if type (cells[c]) == int:
                            try:
                                if run_id is None:
                                    run_id = int (cells[c])
                                    #print (run_id)
                                    if run_id < 1 or run_id > 16:
                                        run_id = None
                            except:
                                pass
                                    
                if p_info is not None and run_id is None:
                    run_id = 1 + len (ngs_records)
                                
                if run_id is not None and p_info is not None:
                    if run_id in ngs_records:
                        print (cells)
                        print ("DUPLICATE RUN_ID %d in %s" % (run_id, file_path), file = sys.stderr)
                        return None
                    else:
                        ngs_records[run_id] = p_info
                
            except:
                raise
    except:
        print ("FAILED reading XLSX %s" % file_path, file = sys.stderr)
            
    return ngs_records if len (ngs_records) > 0 else None   
                    
##########

def handle_SampleKey_xls (file_path, mb_lookup):
    #PID (MM/DD/YYYY) junk
    pid_format       = re.compile ('^([\w\-]+)\s+\(([0-9]{2})/([0-9]{2})/([0-9]{4})\).+$')
  
    try:
        xcl_data =  xlrd.open_workbook(file_path)
        for sn,k in enumerate (xcl_data.sheets()):
            if sn == 0:
                ngs_records = {} 
                for r in range (0,k.nrows): 
                    try:
                        run_id = None
                        p_info = None
                        for j,c in enumerate (k.row (r)):
                            if c.ctype == xlrd.XL_CELL_TEXT:
                               m = pid_format.match (c.value)
                               if m is not None:
                                    try:
                                        pid = mb_lookup[m.group (1)]
                                        date = datetime.datetime (int(m.group (4)), int(m.group (2)), int(m.group(3)))
                                        p_info = (pid,date.strftime("%Y%m%d"))
                                    except KeyError as e:
                                        pass   
                            elif c.ctype == xlrd.XL_CELL_NUMBER and run_id is None:
                                try:
                                    run_id = int (c.value)
                                    if run_id < 1 or run_id > 16:
                                        run_id = None                                
                                except:
                                    pass                             

                        if p_info is not None and run_id is not None:
                            if run_id in ngs_records:
                                print ("DUPLICATE RUN_ID %d in %s" % (run_id, file_path), file = sys.stderr)
                                return None
                            else:
                                ngs_records[run_id] =p_info
                    except:
                        pass
                #print (file_path, ngs_records)
                return ngs_records if len (ngs_records) > 0 else None   
    except:
        print ("FAILED reading XLS %s" % file_path, file = sys.stderr)
    return None
    	
##########
    
def handle_patient_key (file_path, mb_lookup):
    try:
        xcl_data =  xlrd.open_workbook(file_path)
        expected_row0 = ["Run Date","Sample Well","MBN","Sample Date","RNA VL","Notes"]
        for sn,k in enumerate (xcl_data.sheets()):
            if sn == 0:
                row0 = [r.value.rstrip().upper() for r in k.row(0)]
                try:
                    for i,r in enumerate(row0):
                        if expected_row0[i].upper() != r:
                            return None
                
                except:
                    return None
                
                ngs_records = {} 
                for r in range (1,k.nrows):
                    try:
                        run_id = k.row(r)[1]
                        mbn = k.row(r)[2]
                        sample_date = k.row(r)[3]
                        if run_id.ctype == xlrd.XL_CELL_NUMBER and mbn.ctype == xlrd.XL_CELL_TEXT and sample_date.ctype == xlrd.XL_CELL_DATE:
                            sample_date = xlrd.xldate_as_tuple (sample_date.value,1)
                            sample_date = datetime.datetime(sample_date[0],sample_date[1],sample_date[2]).strftime("%Y%m%d")
                            pid = mb_lookup[mbn.value]
                            run_id = int (run_id.value)
                            if run_id in ngs_records:
                                print ("DUPLICATE RUN_ID %d in %s" % (run_id, file_path), file = sys.stderr)
                            else:
                                ngs_records[run_id] = (pid, sample_date)
                    except:
                        pass
                
                return ngs_records if len (ngs_records) > 0 else None   
    except:
        print ("FAILED reading XLS %s" % file_path, file = sys.stderr)
    return None
                    
##########

def add_a_file_to_cache (file_name, patient_info, outpath, paths_done, compartment):

    if file_name in paths_done:
        return None
        
    #print (file_name, patient_info)
    
    files_to_move = None
    for req_ext in [['.fna','.qual'],['.fastq']]:
        file_exists = [os.path.exists(file_name+k) for k in req_ext]
        if False not in file_exists:
            files_to_move = [file_name+k for k in req_ext]
            break
        
    if files_to_move is not None:
        pat_id = patient_info[0]
        sample_date = patient_info[1]
        replicate   = 1
        result_path = os.path.join (outpath, pat_id, sample_date, compartment, str (replicate))
        
        while os.path.exists (result_path):
            replicate += 1
            result_path = os.path.join (outpath, pat_id, sample_date, compartment, str (replicate))
             
        os.makedirs (result_path) 
        
        paths_done [file_name] = result_path
        
        for file in files_to_move:
            name_to_move = os.path.split (file)
            copy_to = os.path.join (result_path, name_to_move[1])
            if not os.path.exists (copy_to):
                os.symlink (file, copy_to)
                print ("%s => %s" % (file, copy_to))
        with open (settings.json, "w") as fh:
            json.dump (paths_done, fh,sort_keys=True, indent=4)

##########

def extract_ngs_info (dir_path, ngs_files):
    if dir_path in ngs_files:
        return ngs_files[dir_path]
    
    for p in ngs_files:
        if os.path.commonprefix((p,dir_path))==dir_path:
            return ngs_files[p]
            
    return None
    
arguments = argparse.ArgumentParser(description='Read filenames.')

arguments.add_argument('-i', '--input', help = 'Scan this directory for NGS files', required = True, type = str)
arguments.add_argument('-l', '--lookup', help = 'Masterbook to AIEDRP PID', required = True, type = argparse.FileType('r'))
arguments.add_argument('-o', '--output', help = 'Send the output to ', required = True, type = str)
arguments.add_argument('-c', '--compartment', help = 'Default sampling compartment ', type = str, default = "BP")
arguments.add_argument('-j', '--json', help = '.json cache file', required = True, type = str)

settings = arguments.parse_args()
	
mb_to_aiedrp = {}

lookupReader = csv.reader(settings.lookup)
next(lookupReader)
for line in lookupReader:
    if len (line[2]):
        mb_to_aiedrp [line[2] if len (line[2]) == 4 else line[2][1:]] = line[1].replace('-','')
    else:
        mb_to_aiedrp [line[0]] = line[0]
        mb_to_aiedrp [line[0].upper()] = line[0]
  

udsReader = csv.reader(settings.input, delimiter = "\t")
    
paths_done = {}
base_path = ""

if os.path.exists(settings.json):
    with open (settings.json, "r") as fh:
        paths_done = json.load (fh)
else:
    paths_done = {}
        
dirs_with_excel_files = {} # dir - excel file name
available_ngs_files   = {}
        
for root, dirs, files in os.walk(settings.input):
    for file in files:
        name, ext = os.path.splitext (file)
        if ext in ('.xls','.xlsx'):
            if root not in dirs_with_excel_files:
                dirs_with_excel_files[root] = [file]
            else:
                dirs_with_excel_files[root].append (file)
        elif ext in ('.fna','.qual','.fastq'):
            if root not in available_ngs_files:
                available_ngs_files[root] = set ()
            available_ngs_files[root].add ((root,name))
            
            
if not os.path.exists (settings.output):
    os.makedirs (settings.output) 

for dir,key_files in dirs_with_excel_files.items():
    #handle varios cases 
    ''' 1 "Patient Key"
        single xls 
        "Run Date","Sample Well","MBN","Sample Date","RNA VL ","Notes" columns in the Excel Document
    '''
    records = None
    for xls in key_files:
        name, ext = os.path.splitext (xls)
        if ext == '.xls' and records is None:
            records = handle_patient_key (os.path.join(dir,xls),mb_to_aiedrp)
            if records is None:
                records = handle_SampleKey_xls (os.path.join(dir,xls),mb_to_aiedrp)
        elif ext == '.xlsx' and records is None:
            records = handle_GeneCore_submission (os.path.join(dir,xls),mb_to_aiedrp)
            if records is None:
                records = handle_454_jr (os.path.join(dir,xls),mb_to_aiedrp)
                #if records is not None:
                #    print (records, extract_ngs_info (dir, available_ngs_files))
    
    
    '''if records is not None:
        print (key_files, records)
    '''  
    if records is None:
        print (dir)
              
    if records is not None:
        ngs_info = extract_ngs_info (dir, available_ngs_files)
        if ngs_info is not None:
            for r,p in records.items():
                prefix = str(r) + "."
                suffix = "RL" + str(r)
                for f in ngs_info:
                    if f[1].startswith (prefix) or f[1].endswith(suffix):
                        add_a_file_to_cache (os.path.join (f[0],f[1]), p, settings.output, paths_done, settings.compartment)
    
    
'''         
for line in udsReader:
    for col_set in [[0,1,2], [3,4,5]]:
        file_name = os.path.normpath(os.path.join(base_path,line[col_set[0]]))
        if file_name not in paths_done:
            try:
                #check for the presence of [.fasta and .qual or a .fastq file]
                
                files_to_move = None
                for req_ext in [['.fna','.qual'],['.fastq']]:
                    file_exists = [os.path.exists(file_name+k) for k in req_ext]
                    if False not in file_exists:
                        files_to_move = [file_name+k for k in req_ext]
                        break
                    
                if files_to_move is not None:
                    pat_id = mb_to_aiedrp[line[col_set[1]].upper()]
                    sample_date = datetime.datetime.strptime(line[col_set[2]],"%Y-%m-%dT%H:%M:%S" ).strftime ("%Y%m%d")
                    replicate   = 1
                    result_path = os.path.join (settings.output, pat_id, sample_date, settings.compartment, str (replicate))
                    while os.path.exists (result_path):
                        replicate += 1
                        result_path = os.path.join (settings.output, pat_id, sample_date, settings.compartment, str (replicate))
                         
                    os.makedirs (result_path) 
                    
                    paths_done [file_name] = result_path
                    
                    for file in files_to_move:
                        name_to_move = os.path.split (file)
                        copy_to = os.path.join (result_path, name_to_move[1])
                        if not os.path.exists (copy_to):
                            os.symlink (file, copy_to)
                            print ("%s => %s" % (file, copy_to))
                    with open (settings.json, "w") as fh:
                        json.dump (paths_done, fh,sort_keys=True, indent=4)

            
                
            except KeyError as e:
                pass
            
'''